import csv, os, dateutil, datetime, glob, re, matplotlib
from openpyxl import load_workbook
import numpy as np
import pylab as plt
from Tkinter import Tk
import Tkinter,tkFileDialog,csv,datetime

font = {'weight' : 'bold',
        'size'   : 12}
matplotlib.rc('font', **font)

zoneList = ['5A','5B','6A','6B']
ZoneOESdata = {}
ZoneOESdates = {}
trimmedZoneOESrawspectra = {}
trimmedZoneOESdates = {}
plotDWZone = {}


elementList = ['Cu','In','Ga','Se','Ar','Na','Mo','Ti','O2','H2']
colorList = ['yellow','orange','grey','maroon','red','black','purple','green','blue','bisque']
OESminList = [323.0, 449.0, 415.0, 470.0, 760.0, 587.0, 378.0, 496.0, 775.0, 654.0] #wavelength minimums for OES integration
OESmaxList = [329.0, 453.0, 419.0, 475.0, 766.0, 591.0, 382.0, 508.0, 779.0, 658.0] #wavelength maxs

'''root = Tkinter.Tk()
root.withdraw()
dirname = tkFileDialog.askdirectory(parent=root,initialdir="/",title='Please select a directory')
dirname = str(dirname)'''

dirname = 'Y:\\Nate\\new MC02 OES program\\backup from MC02 computer\\data\\PC 03-12-15'

os.chdir(dirname)

file = '*signals.csv'
for f in glob.iglob(file):
	OEScsvfile = f


allOESdict = {}

OESCSV = csv.DictReader(open(OEScsvfile),delimiter=',')

for row in OESCSV:
	for column, value in row.iteritems():
		allOESdict.setdefault(column,[]).append(value)
		

OESdates=[]
for eachTime in allOESdict['DateTime']:
	OESdates.append((dateutil.parser.parse(eachTime)-datetime.datetime(1970,1,1)).total_seconds())
	
	
# load raw OES spectra
for zone in zoneList:
	ZfileSearch = '*' + zone + '*'
	for f in glob.iglob(ZfileSearch):
		Zfile = f

	ZOESCSV = csv.reader(open(Zfile), delimiter=',')

	firstRow = True
	ZoneOESdata[zone] = []
	ZoneOESdates[zone] = []
	for row in ZOESCSV:
		if not firstRow:
			ZoneOESdata[zone].append(row[1:])
			ZoneOESdates[zone].append((dateutil.parser.parse(row[0])-datetime.datetime(1970,1,1)).total_seconds())
		else:
			wl = row[1:]
			firstRow = False

		
		
		
# load DW position from export from datasystem
excelFile = '*.xlsx'

for f in glob.iglob(excelFile):
	loadExcelFile = f

dataSystemDates=[]
dataSystemDownWeb=[]

firstRow = True
wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
ws4 = wb.get_sheet_by_name(name="Sheet1")
for row in ws4.iter_rows():
	if row[0].value != None and not firstRow:
		dataSystemDates.append((row[0].value-datetime.datetime(1970,1,1)).total_seconds())
		dataSystemDownWeb.append(float(row[1].value))
	else:
		firstRow = False

		
		
#trim DW
maxDW = 250.0+3.5 #3.5 m to end of zone 6
minDW = 0+2.3 #2.31 m to Z5 entrance
maxDWindex = min(range(len(dataSystemDownWeb)), key=lambda i: abs(dataSystemDownWeb[i]-maxDW))


DWcountDown = maxDWindex
DWstart = dataSystemDownWeb[maxDWindex]
while dataSystemDownWeb[DWcountDown] > minDW:
	DWcountDown-=1

minDWindex = DWcountDown

trimmedDW = dataSystemDownWeb[minDWindex:maxDWindex]
trimmedDates = dataSystemDates[minDWindex:maxDWindex]

trimmedOESdates = []
oesCounters = []
count = -1
for date in OESdates:
	count += 1
	if date > trimmedDates[0] and date < trimmedDates[-1]:
		trimmedOESdates.append(date)
		oesCounters.append(count)

for zone in zoneList:
	trimmedZoneOESdates[zone] = []
	trimmedZoneOESrawspectra[zone] = []
	for count in range(len(ZoneOESdates[zone])):
		if ZoneOESdates[zone][count] > trimmedDates[0] and ZoneOESdates[zone][count] < trimmedDates[-1]:
			trimmedZoneOESdates[zone].append(ZoneOESdates[zone][count])
			trimmedZoneOESrawspectra[zone].append(ZoneOESdata[zone][count])
			
	trimmedZoneOESrawspectra[zone] = np.array(trimmedZoneOESrawspectra[zone])

'''	
#interpolate OES data so it matches data system data points
interpOESdict={}
for key in allOESdict.keys():
	if key != 'DateTime':
		interpOESdict[key] = np.interp(dataSystemDates,OESdates[10:],allOESdict[key][10:])#first few data are empty
for key in interpOESdict:
	interpOESdict[key]=interpOESdict[key][minDWindex:maxDWindex]'''

plotDW = np.interp(trimmedOESdates,trimmedDates,trimmedDW)

for key in allOESdict.keys():
	if key != 'DateTime':
		allOESdict[key] = allOESdict[key][min(oesCounters):max(oesCounters)]
		while len(allOESdict[key]) < len(plotDW):
			plotDW = plotDW[:-1]
			print 'deleted item from plotDW'


plotDWZone['5A'] = plotDW - 2.3
plotDWZone['5B'] = plotDW - 2.6
plotDWZone['6A'] = plotDW - 2.9
plotDWZone['6B'] = plotDW - 3.2




if not os.path.isdir(dirname + '/plots'):
	os.mkdir(dirname + '/plots')
	
os.chdir(dirname + '/plots')

if not os.path.isdir('./signals'):
	os.mkdir('./signals')

if not os.path.isdir('./raw spectra'):
	os.mkdir('./raw spectra')


os.chdir('./signals')

for key in allOESdict.keys():
	if re.search('5A',key):
		plt.scatter(plotDWZone['5A'],allOESdict[key])
	if re.search('5B',key):
		plt.scatter(plotDWZone['5B'],allOESdict[key])
	if re.search('6A',key):
		plt.scatter(plotDWZone['6A'],allOESdict[key])
	if re.search('6B',key):
		plt.scatter(plotDWZone['6B'],allOESdict[key])
	plt.xlim(0,250)
	plt.xlabel('DW')
	plt.ylabel(key + ' OES signal')
	plt.title(key + ' OES signal')
	plt.savefig(key,facecolor='w')
	plt.clf()

'''
os.chdir('../raw spectra')

for zone in zoneList:
	if not os.path.isdir('./Z' + zone):
		os.mkdir('./Z' + zone)
		
	os.chdir('./Z' + zone)
		
	print trimmedZoneOESrawspectra[zone].shape
	for each in range(trimmedZoneOESrawspectra[zone].shape[0]):
		plt.plot(wl,trimmedZoneOESrawspectra[zone][each,:])
		for count in range(len(elementList)):
			plt.axvspan(OESminList[count], OESmaxList[count], label = elementList[count], facecolor = colorList[count], alpha=0.5)
		plt.legend(loc = 'upper left')
		plt.xlabel('wavelength (nm)')
		plt.ylabel('intensity')
		plt.title('Z' + zone + ' OES spectrum at ' + 'DW ' + str(round(plotDWZone[zone][each],1)))
		plt.savefig('DW ' + str(round(plotDWZone[zone][each],1)) + '.png',facecolor = 'w', dpi = 600)
		plt.clf()
	os.chdir('../')'''
	
'''elementList = ['Cu','In','Ga','Se','Ar','Na','Mo','Ti','O2','H2']
OESminList = [323.0, 449.0, 415.0, 470.0, 760.0, 587.0, 378.0, 496.0, 775.0, 654.0] #wavelength minimums for OES integration
OESmaxList = [329.0, 453.0, 419.0, 475.0, 766.0, 591.0, 382.0, 508.0, 779.0, 658.0] #wavelength maxs
OESmaxMins = {}
for each in range(len(elementList)):
	#gets indices of wavelength and intensity arrays for integration
	OESmaxMins[str(elementList[each])+'MIN'], OESmaxMins[str(elementList[each])+'MAX'] = get_WL_indices(OESminList[each], OESmaxList[each])'''
