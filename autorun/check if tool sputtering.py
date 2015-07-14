import subprocess, datetime, re, sys
from openpyxl import load_workbook
sys.path.append('Y:/Nate/git/nuvosun-python-lib')
import nuvosunlib as nsl

print 'checking schedule file'
monitorStartFile = 'C:/OESdata/monitorProcessStart.py'
tool = sys.argv[1]

today = datetime.datetime.now()
    
schedF = nsl.getLatestScheduleFile()

wb = load_workbook(schedF, data_only=True, read_only=True)
whiteBoardWS = wb.get_sheet_by_name("White Board")
rowCount = 0
for row in whiteBoardWS.iter_rows():
    if row[0].value == 'MC01':
        MC01row = rowCount
    elif row[0].value == 'MC02':
        MC02row = rowCount
    elif row[0].value == 'MC03':
        MC03row = rowCount
    elif row[0].value == 'LC1':
        LC1row = rowCount
    rowCount +=1
if tool == 'MC01':
    toolRow = MC01row
elif tool == 'MC02':
    toolRow = MC02row
for column in whiteBoardWS.columns:
    try:
        columnDate = column[1].value
        if columnDate.year == today.year and columnDate.month == today.month and columnDate.day == today.day:
            print 'matched date in worksheet,', columnDate.year, columnDate.month, columnDate.day
            print column[toolRow].value
            if column[toolRow].value == None:
                print 'starting monitor for BE or PC start'
                subprocess.Popen(['python',monitorStartFile,'PC+BE',tool])
            elif re.search('BE', column[toolRow].value) and not re.search('PM', column[toolRow].value):
                runNum = re.search('\d\d\d', column[toolRow].value).group(0)
                print 'starting monitor for BE start, run', runNum
                subprocess.Popen(['python',monitorStartFile,'BE',tool,runNum])
            elif re.search('PC', column[toolRow].value) and not re.search('PM', column[toolRow].value):
                runNum = re.search('\d\d\d', column[toolRow].value).group(0)
                print 'starting monitor for PC start, run', runNum
                subprocess.Popen(['python',monitorStartFile,'PC',tool,runNum])
            elif re.search('PM', column[toolRow].value):
                print 'found PM, exiting'
                exit()
    except Exception as e:
        print e